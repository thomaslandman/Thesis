import os
from shutil import copyfile
import pandas as pd

def mkdir(dir):
    if not os.path.exists(dir):
        os.makedirs(dir)

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

exp_superfolder = "../experiments"  # <<-----x Experiments Superfolder
template_file = {'HMC':'./Evaluation-HMC-Template.xlsx',
                 'EMC':'./Evaluation-EMC-Template.xlsx'}  # <<-----x Experiments Superfolder
all_xlsx_files_dir = './all_xlsx_files/'
if not os.path.exists (all_xlsx_files_dir):
    os.makedirs(all_xlsx_files_dir)

datasets = ['HMC', 'EMC']
num_patients = {'HMC':50, 'EMC':42}
tasks_list = os.listdir(exp_superfolder)

for task in tasks_list:
    exp_folder_list = [x for x in os.listdir(os.path.join(exp_superfolder, task)) if
                       os.path.isdir(os.path.join(exp_superfolder, task, x))]
    for exp_folder in exp_folder_list:
        for ds in datasets:
            exp_xls_dir = os.path.join(exp_superfolder, task, exp_folder, "output", ds)
            if os.path.exists(exp_xls_dir):
                xls_file_list = [x for x in os.listdir(exp_xls_dir) if ".xlsx" in x]
                for xls_file in xls_file_list:
                    file_parts = xls_file.split(".")
                    src_file = os.path.join(exp_xls_dir, xls_file)
                    mkdir(os.path.join(all_xlsx_files_dir, ds))
                    dst_file = os.path.join(all_xlsx_files_dir, ds, file_parts[0] + "-" + task + "-" + exp_folder + "." + file_parts[1])
                    copyfile(template_file[ds], dst_file)
                    data = pd.read_excel(src_file, engine='openpyxl', sheet_name="eval").iloc[:num_patients[ds], 1:]
                    append_df_to_excel(dst_file, data, sheet_name='eval', startcol=0, startrow=0)
                    print("COPYING " + src_file + " to " + dst_file)
